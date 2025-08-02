import psycopg2
import logging
import os
from datetime import datetime
from dotenv import load_dotenv
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import requests

# Configuração do Logging
LOG_FILE = "erros.log"
logging.basicConfig(filename=LOG_FILE, level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# Carregar variáveis de ambiente
load_dotenv()

class DataService:
    """Gerencia todas as interações com o banco de dados PostgreSQL e a lógica de negócios."""
    def __init__(self):
        self.db_params = {
            "dbname": os.getenv("DB_NAME"),
            "user": os.getenv("DB_USER"),
            "password": os.getenv("DB_PASSWORD"),
            "host": os.getenv("DB_HOST"),
            "port": os.getenv("DB_PORT")
        }
        self.init_db()

    def get_connection(self):
        """Cria uma conexão com o PostgreSQL."""
        try:
            return psycopg2.connect(**self.db_params)
        except psycopg2.Error as e:
            logging.error(f"Erro ao conectar ao banco de dados: {e}")
            raise Exception(f"Não foi possível conectar ao banco: {e}")

    def init_db(self):
        """Inicializa o banco de dados PostgreSQL com a estrutura necessária."""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as c:
                    # Criar tabelas, se não existirem
                    c.execute('''
                        CREATE TABLE IF NOT EXISTS planos_internos (
                            codigo TEXT PRIMARY KEY
                        );
                    ''')
                    c.execute('''
                        CREATE TABLE IF NOT EXISTS naturezas_despesa (
                            codigo TEXT PRIMARY KEY,
                            plano_interno_codigo TEXT,
                            FOREIGN KEY (plano_interno_codigo) REFERENCES planos_internos (codigo) ON DELETE CASCADE
                        );
                    ''')
                    c.execute('''
                        CREATE TABLE IF NOT EXISTS secoes_requisitantes (
                            codigo TEXT PRIMARY KEY
                        );
                    ''')
                    c.execute('''
                        CREATE TABLE IF NOT EXISTS notas (
                            numero TEXT PRIMARY KEY,
                            valor REAL,
                            valor_restante REAL,
                            descricao TEXT,
                            observacao TEXT,
                            prazo TEXT,
                            data_criacao TEXT,
                            natureza_despesa_codigo TEXT,
                            plano_interno_codigo TEXT,
                            ptres_codigo TEXT,
                            fonte_codigo TEXT,
                            FOREIGN KEY (natureza_despesa_codigo) REFERENCES naturezas_despesa (codigo) ON DELETE CASCADE,
                            FOREIGN KEY (plano_interno_codigo) REFERENCES planos_internos (codigo) ON DELETE CASCADE
                        );
                    ''')
                    c.execute('''
                        CREATE TABLE IF NOT EXISTS empenhos (
                            id SERIAL PRIMARY KEY,
                            numero_nota TEXT,
                            valor REAL,
                            descricao TEXT,
                            data TEXT,
                            secao_requisitante_codigo TEXT,
                            FOREIGN KEY (numero_nota) REFERENCES notas (numero) ON DELETE CASCADE,
                            FOREIGN KEY (secao_requisitante_codigo) REFERENCES secoes_requisitantes (codigo) ON DELETE SET NULL
                        );
                    ''')
                    conn.commit()
                    logging.info("Banco de dados inicializado com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao inicializar banco de dados: {e}")
            raise Exception(f"Não foi possível inicializar o banco: {e}")

    def load_data(self, query, params=None):
        """Carrega dados do banco de dados com segurança."""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as c:
                    if params:
                        c.execute(query, params)
                    else:
                        c.execute(query)
                    return c.fetchall()
        except Exception as e:
            logging.error(f"Erro ao carregar dados ({query}): {e}")
            raise Exception(f"Não foi possível ler os dados: {e}")

    def load_planos_internos(self):
        """Carrega todos os planos internos."""
        rows = self.load_data("SELECT codigo FROM planos_internos ORDER BY codigo")
        return [{"codigo": r[0]} for r in rows]

    def load_naturezas_despesa(self, plano_interno_codigo=None):
        """Carrega naturezas de despesa, opcionalmente filtrando por plano interno."""
        query = "SELECT codigo, plano_interno_codigo FROM naturezas_despesa"
        params = None
        if plano_interno_codigo:
            query += " WHERE plano_interno_codigo = %s"
            params = (plano_interno_codigo,)
        rows = self.load_data(query, params)
        return [{"codigo": r[0], "plano_interno_codigo": r[1]} for r in rows]

    def load_secoes_requisitantes(self):
        """Carrega todas as seções requisitantes."""
        rows = self.load_data("SELECT codigo FROM secoes_requisitantes ORDER BY codigo")
        return [{"codigo": r[0]} for r in rows]

    def load_notas(self, natureza_despesa_codigo=None):
        """Carrega todas as notas, opcionalmente filtrando por natureza da despesa."""
        query = "SELECT numero, valor, valor_restante, descricao, observacao, prazo, data_criacao, natureza_despesa_codigo, plano_interno_codigo, ptres_codigo, fonte_codigo FROM notas ORDER BY data_criacao DESC"
        params = None
        if natureza_despesa_codigo:
            query = query.replace("ORDER BY", "WHERE natureza_despesa_codigo = %s ORDER BY")
            params = (natureza_despesa_codigo,)
        rows = self.load_data(query, params)
        return [{
            "numero": r[0], "valor": r[1], "valor_restante": r[2], "descricao": r[3],
            "observacao": r[4], "prazo": r[5], "data_criacao": r[6],
            "natureza_despesa_codigo": r[7], "plano_interno_codigo": r[8],
            "ptres_codigo": r[9], "fonte_codigo": r[10]
        } for r in rows]

    def load_empenhos(self, numero_nota=None):
        """Carrega empenhos, opcionalmente filtrando por número de nota."""
        query = "SELECT id, numero_nota, valor, descricao, data, secao_requisitante_codigo FROM empenhos"
        params = None
        if numero_nota:
            query += " WHERE numero_nota = %s"
            params = (numero_nota,)
        rows = self.load_data(query, params)
        return [{
            "id": r[0], "numero_nota": r[1], "valor": r[2], "descricao": r[3],
            "data": r[4], "secao_requisitante_codigo": r[5]
        } for r in rows]

    def save_plano_interno(self, plano_interno):
        """Salva um novo plano interno."""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as c:
                    c.execute("INSERT INTO planos_internos (codigo) VALUES (%s)",
                              (plano_interno["codigo"],))
                    conn.commit()
            logging.info(f"Plano Interno {plano_interno['codigo']} salvo com sucesso.")
        except psycopg2.IntegrityError:
            logging.error(f"Erro de integridade ao salvar plano interno: {plano_interno['codigo']} já existe.")
            raise ValueError(f"O código do plano interno '{plano_interno['codigo']}' já existe.")
        except Exception as e:
            logging.error(f"Erro ao salvar plano interno: {e}")
            raise

    def save_natureza_despesa(self, natureza_despesa):
        """Salva uma nova natureza da despesa."""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as c:
                    c.execute("INSERT INTO naturezas_despesa (codigo, plano_interno_codigo) VALUES (%s, %s)",
                              (natureza_despesa["codigo"], natureza_despesa["plano_interno_codigo"]))
                    conn.commit()
            logging.info(f"Natureza da Despesa {natureza_despesa['codigo']} salva com sucesso.")
        except psycopg2.IntegrityError:
            logging.error(f"Erro de integridade ao salvar natureza da despesa: {natureza_despesa['codigo']} já existe.")
            raise ValueError(f"O código da natureza da despesa '{natureza_despesa['codigo']}' já existe.")
        except Exception as e:
            logging.error(f"Erro ao salvar natureza da despesa: {e}")
            raise

    def save_secao_requisitante(self, secao_requisitante):
        """Salva uma nova seção requisitante."""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as c:
                    c.execute("INSERT INTO secoes_requisitantes (codigo) VALUES (%s)",
                              (secao_requisitante["codigo"],))
                    conn.commit()
            logging.info(f"Seção Requisitante {secao_requisitante['codigo']} salva com sucesso.")
        except psycopg2.IntegrityError:
            logging.error(f"Erro de integridade ao salvar seção requisitante: {secao_requisitante['codigo']} já existe.")
            raise ValueError(f"O código da seção requisitante '{secao_requisitante['codigo']}' já existe.")
        except Exception as e:
            logging.error(f"Erro ao salvar seção requisitante: {e}")
            raise

    def save_nota(self, nota):
        """Salva uma nova nota no banco de dados."""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as c:
                    c.execute("""
                        INSERT INTO notas (numero, valor, valor_restante, descricao, observacao, prazo, data_criacao,
                                           natureza_despesa_codigo, plano_interno_codigo, ptres_codigo, fonte_codigo)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        nota["numero"], nota["valor"], nota["valor_restante"], nota["descricao"],
                        nota["observacao"], nota["prazo"], nota["data_criacao"],
                        nota["natureza_despesa_codigo"], nota["plano_interno_codigo"],
                        nota["ptres_codigo"], nota["fonte_codigo"]
                    ))
                    conn.commit()
            logging.info(f"Nota {nota['numero']} salva com sucesso.")
        except psycopg2.IntegrityError:
            logging.error(f"Erro de integridade ao salvar nota: {nota['numero']} já existe.")
            raise ValueError(f"O número de nota '{nota['numero']}' já existe.")
        except Exception as e:
            logging.error(f"Erro ao salvar nota: {e}")
            raise

    def save_empenho(self, empenho, nota):
        """Salva um empenho e atualiza o valor restante da nota."""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as c:
                    c.execute("""
                        INSERT INTO empenhos (numero_nota, valor, descricao, data, secao_requisitante_codigo)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (
                        empenho["numero_nota"], empenho["valor"], empenho["descricao"],
                        empenho["data"], empenho["secao_requisitante_codigo"]
                    ))
                    c.execute("UPDATE notas SET valor_restante = %s WHERE numero = %s",
                              (nota["valor_restante"], nota["numero"]))
                    conn.commit()
            logging.info(f"Empenho de R${empenho['valor']:.2f} registrado para nota {empenho['numero_nota']}.")
        except Exception as e:
            logging.error(f"Erro ao salvar empenho: {e}")
            raise

    def delete_plano_interno(self, codigo):
        """Deleta um plano interno."""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as c:
                    c.execute("DELETE FROM planos_internos WHERE codigo = %s", (codigo,))
                    conn.commit()
            logging.info(f"Plano Interno {codigo} deletado com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao deletar plano interno {codigo}: {e}")
            raise

    def delete_natureza_despesa(self, codigo):
        """Deleta uma natureza da despesa."""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as c:
                    c.execute("DELETE FROM naturezas_despesa WHERE codigo = %s", (codigo,))
                    conn.commit()
            logging.info(f"Natureza da Despesa {codigo} deletada com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao deletar natureza da despesa {codigo}: {e}")
            raise

    def delete_secao_requisitante(self, codigo):
        """Deleta uma seção requisitante."""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as c:
                    c.execute("DELETE FROM secoes_requisitantes WHERE codigo = %s", (codigo,))
                    conn.commit()
            logging.info(f"Seção Requisitante {codigo} deletada com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao deletar seção requisitante {codigo}: {e}")
            raise

    def delete_nota(self, numero_nota):
        """Deleta uma nota do banco de dados."""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as c:
                    c.execute("DELETE FROM notas WHERE numero = %s", (numero_nota,))
                    conn.commit()
            logging.info(f"Nota {numero_nota} deletada com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao deletar nota {numero_nota}: {e}")
            raise

    def delete_empenho(self, empenho_id):
        """Deleta um empenho e atualiza o valor restante da nota associada."""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as c:
                    c.execute("SELECT numero_nota, valor FROM empenhos WHERE id = %s", (empenho_id,))
                    empenho = c.fetchone()
                    if not empenho:
                        raise ValueError("Empenho não encontrado.")
                    numero_nota, valor_empenho = empenho
                    c.execute("UPDATE notas SET valor_restante = valor_restante + %s WHERE numero = %s",
                              (valor_empenho, numero_nota))
                    c.execute("DELETE FROM empenhos WHERE id = %s", (empenho_id,))
                    conn.commit()
            logging.info(f"Empenho {empenho_id} deletado com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao deletar empenho {empenho_id}: {e}")
            raise

    def generate_excel_report(self):
        """Gera relatório em Excel com uma linha por empenho, incluindo a hierarquia."""
        notas = self.load_notas()
        empenhos = self.load_empenhos()
        naturezas = self.load_naturezas_despesa()
        planos = self.load_planos_internos()
        secoes = self.load_secoes_requisitantes()

        total_valor_geral = sum(n["valor"] for n in notas)
        total_restante_geral = sum(n["valor_restante"] for n in notas)
        total_empenhado_geral = sum(e["valor"] for e in empenhos)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório Detalhado de Empenhos"

        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

        headers = [
            "Plano Interno", "Natureza da Despesa", "PTRES", "Fonte",
            "Nº Nota", "Valor Original (R$)", "Valor Restante (R$)", "Descrição da Nota", "Prazo",
            "Data Empenho", "Valor Empenho (R$)", "Descrição Empenho", "Seção Requisitante"
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border

        for nota in notas:
            natureza = next((n for n in naturezas if n["codigo"] == nota["natureza_despesa_codigo"]), {"codigo": "N/A", "plano_interno_codigo": None})
            plano = next((p for p in planos if p["codigo"] == natureza["plano_interno_codigo"]), {"codigo": "N/A"}) if natureza["plano_interno_codigo"] else {"codigo": "N/A"}

            related_empenhos = [e for e in empenhos if e["numero_nota"] == nota["numero"]]
            if not related_empenhos:
                row_data = [
                    plano["codigo"], natureza["codigo"], nota["ptres_codigo"], nota["fonte_codigo"],
                    nota["numero"], nota["valor"], nota["valor_restante"], nota["descricao"], nota["prazo"],
                    "Nenhum empenho", "", "", "N/A"
                ]
                ws.append(row_data)
            else:
                for empenho in related_empenhos:
                    secao = next((s for s in secoes if s["codigo"] == empenho["secao_requisitante_codigo"]), {"codigo": "N/A"}) if empenho["secao_requisitante_codigo"] else {"codigo": "N/A"}
                    row_data = [
                        plano["codigo"], natureza["codigo"], nota["ptres_codigo"], nota["fonte_codigo"],
                        nota["numero"], nota["valor"], nota["valor_restante"], nota["descricao"], nota["prazo"],
                        empenho["data"], empenho["valor"], empenho["descricao"], secao["codigo"]
                    ]
                    ws.append(row_data)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = align_left if cell.column in [1, 2, 4, 8, 12, 13] else align_center

        ws.append([])  # Linha em branco
        total_row_idx = ws.max_row + 1
        ws.cell(row=total_row_idx, column=1, value="TOTAIS GERAIS").font = header_font
        ws.cell(row=total_row_idx, column=6, value=total_valor_geral).font = Font(bold=True)
        ws.cell(row=total_row_idx, column=7, value=total_restante_geral).font = Font(bold=True)
        ws.cell(row=total_row_idx, column=11, value=total_empenhado_geral).font = Font(bold=True)

        for col in range(1, len(headers) + 1):
            ws.cell(row=total_row_idx, column=col).border = border
            ws.cell(row=total_row_idx, column=col).alignment = align_center
        ws.cell(row=total_row_idx, column=1).fill = header_fill

        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col[0].column_letter].width = min(adjusted_width, 50)

        filename = "relatorios/relatorio_notas_credito.xlsx"
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        wb.save(filename)
        logging.info("Relatório Excel gerado com sucesso.")
        return filename

    def generate_pdf_report(self):
        """Gera relatório em PDF com uma linha por empenho, incluindo a hierarquia."""
        notas = self.load_notas()
        empenhos = self.load_empenhos()
        naturezas = self.load_naturezas_despesa()
        planos = self.load_planos_internos()
        secoes = self.load_secoes_requisitantes()

        total_valor_geral = sum(n["valor"] for n in notas)
        total_restante_geral = sum(n["valor_restante"] for n in notas)
        total_empenhado_geral = sum(e["valor"] for e in empenhos)

        filename = "relatorios/relatorio_notas_credito.pdf"
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        doc = SimpleDocTemplate(filename, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = []

        headers = [
            "Plano Interno", "Natureza da Despesa", "PTRES", "Fonte",
            "Nº Nota", "V. Original", "V. Restante", "Data Empenho", "V. Empenho", "Descrição Empenho", "Seção Requisitante"
        ]
        data = [headers]

        for nota in notas:
            natureza = next((n for n in naturezas if n["codigo"] == nota["natureza_despesa_codigo"]), {"codigo": "N/A", "plano_interno_codigo": None})
            plano = next((p for p in planos if p["codigo"] == natureza["plano_interno_codigo"]), {"codigo": "N/A"}) if natureza["plano_interno_codigo"] else {"codigo": "N/A"}

            related_empenhos = [e for e in empenhos if e["numero_nota"] == nota["numero"]]
            if not related_empenhos:
                data.append([
                    Paragraph(plano["codigo"], styles['BodyText']),
                    Paragraph(natureza["codigo"], styles['BodyText']),
                    Paragraph(nota["ptres_codigo"], styles['BodyText']),
                    Paragraph(nota["fonte_codigo"], styles['BodyText']),
                    Paragraph(nota["numero"], styles['BodyText']),
                    f"R$ {nota['valor']:.2f}", f"R$ {nota['valor_restante']:.2f}",
                    "Nenhum", "", Paragraph(nota["descricao"], styles['BodyText']), "N/A"
                ])
            else:
                for empenho in related_empenhos:
                    secao = next((s for s in secoes if s["codigo"] == empenho["secao_requisitante_codigo"]), {"codigo": "N/A"}) if empenho["secao_requisitante_codigo"] else {"codigo": "N/A"}
                    data.append([
                        Paragraph(plano["codigo"], styles['BodyText']),
                        Paragraph(natureza["codigo"], styles['BodyText']),
                        Paragraph(nota["ptres_codigo"], styles['BodyText']),
                        Paragraph(nota["fonte_codigo"], styles['BodyText']),
                        Paragraph(nota["numero"], styles['BodyText']),
                        f"R$ {nota['valor']:.2f}", f"R$ {nota['valor_restante']:.2f}",
                        empenho["data"], f"R$ {empenho['valor']:.2f}", Paragraph(empenho["descricao"], styles['BodyText']),
                        Paragraph(secao["codigo"], styles['BodyText'])
                    ])

        table = Table(data, colWidths=[80, 80, 80, 80, 80, 80, 80, 100, 80, 180, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E90FF")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ]))

        title = Paragraph("Relatório Detalhado de Notas de Crédito e Empenhos", styles['h1'])
        elements.append(title)
        elements.append(table)
        elements.append(Spacer(1, 20))

        total_data = [
            ["Totais Gerais", "", "", "", "Valor Original Total", "Valor Restante Total", "", "Valor Empenhado Total", "", ""],
            ["", "", "", "", f"R$ {total_valor_geral:.2f}", f"R$ {total_restante_geral:.2f}", "", f"R$ {total_empenhado_geral:.2f}", "", ""]
        ]
        total_table = Table(total_data, colWidths=[80, 80, 80, 80, 80, 80, 100, 80, 180, 80])
        total_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(total_table)

        doc.build(elements)
        logging.info("Relatório PDF gerado com sucesso.")
        return filename

    def consultar_siafi(self, usuario, senha, api_key=None):
        """Consulta o SIAFI com autenticação (simulada para testes)."""
        logging.info("Consulta SIAFI simulada para testes.")
        return {"status": "success", "message": "Consulta simulada"}