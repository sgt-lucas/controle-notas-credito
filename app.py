import streamlit as st
import pandas as pd
import re
from datetime import datetime
import io
import os
import logging
import psycopg2
from psycopg2 import pool
from dotenv import load_dotenv
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Configuração do Logging
LOG_FILE = "erros.log"
logging.basicConfig(filename=LOG_FILE, level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# Carregar variáveis de ambiente
load_dotenv()

class DataService:
    """Gerencia todas as interações com o banco de dados PostgreSQL e a lógica de negócios."""
    def __init__(self):
        self.db_params = {
            "dbname": os.getenv("DB_NAME"),
            "user": os.getenv("DB_USER"),
            "password": os.getenv("DB_PASSWORD"),
            "host": os.getenv("DB_HOST"),
            "port": os.getenv("DB_PORT")
        }
        # Configurar pool de conexões para suportar até 3 usuários
        self.connection_pool = psycopg2.pool.ThreadedConnectionPool(
            minconn=1, maxconn=3, **self.db_params
        )
        self.init_db()

    def get_connection(self):
        """Obtém uma conexão do pool."""
        try:
            return self.connection_pool.getconn()
        except psycopg2.Error as e:
            logging.error(f"Erro ao obter conexão do pool: {e}")
            raise Exception(f"Não foi possível conectar ao banco: {e}")

    def release_connection(self, conn):
        """Libera a conexão de volta ao pool."""
        self.connection_pool.putconn(conn)

    def init_db(self):
        """Inicializa o banco de dados PostgreSQL com a estrutura necessária."""
        conn = self.get_connection()
        try:
            with conn.cursor() as c:
                c.execute('''
                    CREATE TABLE IF NOT EXISTS planos_internos (
                        codigo TEXT PRIMARY KEY
                    );
                ''')
                c.execute('''
                    CREATE TABLE IF NOT EXISTS naturezas_despesa (
                        codigo TEXT PRIMARY KEY,
                        plano_interno_codigo TEXT,
                        FOREIGN KEY (plano_interno_codigo) REFERENCES planos_internos (codigo) ON DELETE CASCADE
                    );
                ''')
                c.execute('''
                    CREATE TABLE IF NOT EXISTS secoes_requisitantes (
                        codigo TEXT PRIMARY KEY
                    );
                ''')
                c.execute('''
                    CREATE TABLE IF NOT EXISTS notas (
                        numero TEXT PRIMARY KEY,
                        valor REAL,
                        valor_restante REAL,
                        descricao TEXT,
                        observacao TEXT,
                        prazo TEXT,
                        data_criacao TEXT,
                        natureza_despesa_codigo TEXT,
                        plano_interno_codigo TEXT,
                        ptres_codigo TEXT,
                        fonte_codigo TEXT,
                        FOREIGN KEY (natureza_despesa_codigo) REFERENCES naturezas_despesa (codigo) ON DELETE CASCADE,
                        FOREIGN KEY (plano_interno_codigo) REFERENCES planos_internos (codigo) ON DELETE CASCADE
                    );
                ''')
                c.execute('''
                    CREATE TABLE IF NOT EXISTS empenhos (
                        id SERIAL PRIMARY KEY,
                        numero_nota TEXT,
                        valor REAL,
                        descricao TEXT,
                        data TEXT,
                        secao_requisitante_codigo TEXT,
                        FOREIGN KEY (numero_nota) REFERENCES notas (numero) ON DELETE CASCADE,
                        FOREIGN KEY (secao_requisitante_codigo) REFERENCES secoes_requisitantes (codigo) ON DELETE SET NULL
                    );
                ''')
                conn.commit()
                logging.info("Banco de dados inicializado com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao inicializar banco de dados: {e}")
            raise Exception(f"Não foi possível inicializar o banco: {e}")
        finally:
            self.release_connection(conn)

    def load_data(self, query, params=None):
        """Carrega dados do banco de dados com segurança."""
        conn = self.get_connection()
        try:
            with conn.cursor() as c:
                if params:
                    c.execute(query, params)
                else:
                    c.execute(query)
                return c.fetchall()
        except Exception as e:
            logging.error(f"Erro ao carregar dados ({query}): {e}")
            raise Exception(f"Não foi possível ler os dados: {e}")
        finally:
            self.release_connection(conn)

    def load_planos_internos(self):
        """Carrega todos os planos internos."""
        rows = self.load_data("SELECT codigo FROM planos_internos ORDER BY codigo")
        return [{"codigo": r[0]} for r in rows]

    def load_naturezas_despesa(self, plano_interno_codigo=None):
        """Carrega naturezas de despesa, opcionalmente filtrando por plano interno."""
        query = "SELECT codigo, plano_interno_codigo FROM naturezas_despesa"
        params = None
        if plano_interno_codigo:
            query += " WHERE plano_interno_codigo = %s"
            params = (plano_interno_codigo,)
        rows = self.load_data(query, params)
        return [{"codigo": r[0], "plano_interno_codigo": r[1]} for r in rows]

    def load_secoes_requisitantes(self):
        """Carrega todas as seções requisitantes."""
        rows = self.load_data("SELECT codigo FROM secoes_requisitantes ORDER BY codigo")
        return [{"codigo": r[0]} for r in rows]

    def load_notas(self, natureza_despesa_codigo=None):
        """Carrega todas as notas, opcionalmente filtrando por natureza da despesa."""
        query = "SELECT numero, valor, valor_restante, descricao, observacao, prazo, data_criacao, natureza_despesa_codigo, plano_interno_codigo, ptres_codigo, fonte_codigo FROM notas ORDER BY data_criacao DESC"
        params = None
        if natureza_despesa_codigo:
            query = query.replace("ORDER BY", "WHERE natureza_despesa_codigo = %s ORDER BY")
            params = (natureza_despesa_codigo,)
        rows = self.load_data(query, params)
        return [{
            "numero": r[0], "valor": r[1], "valor_restante": r[2], "descricao": r[3],
            "observacao": r[4], "prazo": r[5], "data_criacao": r[6],
            "natureza_despesa_codigo": r[7], "plano_interno_codigo": r[8],
            "ptres_codigo": r[9], "fonte_codigo": r[10]
        } for r in rows]

    def load_empenhos(self, numero_nota=None):
        """Carrega empenhos, opcionalmente filtrando por número de nota."""
        query = "SELECT id, numero_nota, valor, descricao, data, secao_requisitante_codigo FROM empenhos"
        params = None
        if numero_nota:
            query += " WHERE numero_nota = %s"
            params = (numero_nota,)
        rows = self.load_data(query, params)
        return [{
            "id": r[0], "numero_nota": r[1], "valor": r[2], "descricao": r[3],
            "data": r[4], "secao_requisitante_codigo": r[5]
        } for r in rows]

    def save_plano_interno(self, plano_interno):
        """Salva um novo plano interno."""
        conn = self.get_connection()
        try:
            with conn.cursor() as c:
                c.execute("INSERT INTO planos_internos (codigo) VALUES (%s)",
                          (plano_interno["codigo"],))
                conn.commit()
            logging.info(f"Plano Interno {plano_interno['codigo']} salvo com sucesso.")
        except psycopg2.IntegrityError:
            logging.error(f"Erro de integridade ao salvar plano interno: {plano_interno['codigo']} já existe.")
            raise ValueError(f"O código do plano interno '{plano_interno['codigo']}' já existe.")
        except Exception as e:
            logging.error(f"Erro ao salvar plano interno: {e}")
            raise
        finally:
            self.release_connection(conn)

    def save_natureza_despesa(self, natureza_despesa):
        """Salva uma nova natureza da despesa."""
        conn = self.get_connection()
        try:
            with conn.cursor() as c:
                c.execute("INSERT INTO naturezas_despesa (codigo, plano_interno_codigo) VALUES (%s, %s)",
                          (natureza_despesa["codigo"], natureza_despesa["plano_interno_codigo"]))
                conn.commit()
            logging.info(f"Natureza da Despesa {natureza_despesa['codigo']} salva com sucesso.")
        except psycopg2.IntegrityError:
            logging.error(f"Erro de integridade ao salvar natureza da despesa: {natureza_despesa['codigo']} já existe.")
            raise ValueError(f"O código da natureza da despesa '{natureza_despesa['codigo']}' já existe.")
        except Exception as e:
            logging.error(f"Erro ao salvar natureza da despesa: {e}")
            raise
        finally:
            self.release_connection(conn)

    def save_secao_requisitante(self, secao_requisitante):
        """Salva uma nova seção requisitante."""
        conn = self.get_connection()
        try:
            with conn.cursor() as c:
                c.execute("INSERT INTO secoes_requisitantes (codigo) VALUES (%s)",
                          (secao_requisitante["codigo"],))
                conn.commit()
            logging.info(f"Seção Requisitante {secao_requisitante['codigo']} salva com sucesso.")
        except psycopg2.IntegrityError:
            logging.error(f"Erro de integridade ao salvar seção requisitante: {secao_requisitante['codigo']} já existe.")
            raise ValueError(f"O código da seção requisitante '{secao_requisitante['codigo']}' já existe.")
        except Exception as e:
            logging.error(f"Erro ao salvar seção requisitante: {e}")
            raise
        finally:
            self.release_connection(conn)

    def save_nota(self, nota):
        """Salva uma nova nota no banco de dados."""
        conn = self.get_connection()
        try:
            with conn.cursor() as c:
                c.execute("""
                    INSERT INTO notas (numero, valor, valor_restante, descricao, observacao, prazo, data_criacao,
                                       natureza_despesa_codigo, plano_interno_codigo, ptres_codigo, fonte_codigo)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    nota["numero"], nota["valor"], nota["valor_restante"], nota["descricao"],
                    nota["observacao"], nota["prazo"], nota["data_criacao"],
                    nota["natureza_despesa_codigo"], nota["plano_interno_codigo"],
                    nota["ptres_codigo"], nota["fonte_codigo"]
                ))
                conn.commit()
            logging.info(f"Nota {nota['numero']} salva com sucesso.")
        except psycopg2.IntegrityError:
            logging.error(f"Erro de integridade ao salvar nota: {nota['numero']} já existe.")
            raise ValueError(f"O número de nota '{nota['numero']}' já existe.")
        except Exception as e:
            logging.error(f"Erro ao salvar nota: {e}")
            raise
        finally:
            self.release_connection(conn)

    def save_empenho(self, empenho, nota):
        """Salva um empenho e atualiza o valor restante da nota."""
        conn = self.get_connection()
        try:
            with conn.cursor() as c:
                c.execute("""
                    INSERT INTO empenhos (numero_nota, valor, descricao, data, secao_requisitante_codigo)
                    VALUES (%s, %s, %s, %s, %s)
                """, (
                    empenho["numero_nota"], empenho["valor"], empenho["descricao"],
                    empenho["data"], empenho["secao_requisitante_codigo"]
                ))
                c.execute("UPDATE notas SET valor_restante = %s WHERE numero = %s",
                          (nota["valor_restante"], nota["numero"]))
                conn.commit()
            logging.info(f"Empenho de R${empenho['valor']:.2f} registrado para nota {empenho['numero_nota']}.")
        except Exception as e:
            logging.error(f"Erro ao salvar empenho: {e}")
            raise
        finally:
            self.release_connection(conn)

    def delete_plano_interno(self, codigo):
        """Deleta um plano interno."""
        conn = self.get_connection()
        try:
            with conn.cursor() as c:
                c.execute("DELETE FROM planos_internos WHERE codigo = %s", (codigo,))
                conn.commit()
            logging.info(f"Plano Interno {codigo} deletado com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao deletar plano interno {codigo}: {e}")
            raise
        finally:
            self.release_connection(conn)

    def delete_natureza_despesa(self, codigo):
        """Deleta uma natureza da despesa."""
        conn = self.get_connection()
        try:
            with conn.cursor() as c:
                c.execute("DELETE FROM naturezas_despesa WHERE codigo = %s", (codigo,))
                conn.commit()
            logging.info(f"Natureza da Despesa {codigo} deletada com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao deletar natureza da despesa {codigo}: {e}")
            raise
        finally:
            self.release_connection(conn)

    def delete_secao_requisitante(self, codigo):
        """Deleta uma seção requisitante."""
        conn = self.get_connection()
        try:
            with conn.cursor() as c:
                c.execute("DELETE FROM secoes_requisitantes WHERE codigo = %s", (codigo,))
                conn.commit()
            logging.info(f"Seção Requisitante {codigo} deletada com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao deletar seção requisitante {codigo}: {e}")
            raise
        finally:
            self.release_connection(conn)

    def delete_nota(self, numero_nota):
        """Deleta uma nota do banco de dados."""
        conn = self.get_connection()
        try:
            with conn.cursor() as c:
                c.execute("DELETE FROM notas WHERE numero = %s", (numero_nota,))
                conn.commit()
            logging.info(f"Nota {numero_nota} deletada com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao deletar nota {numero_nota}: {e}")
            raise
        finally:
            self.release_connection(conn)

    def delete_empenho(self, empenho_id):
        """Deleta um empenho e atualiza o valor restante da nota associada."""
        conn = self.get_connection()
        try:
            with conn.cursor() as c:
                c.execute("SELECT numero_nota, valor FROM empenhos WHERE id = %s", (empenho_id,))
                empenho = c.fetchone()
                if not empenho:
                    raise ValueError("Empenho não encontrado.")
                numero_nota, valor_empenho = empenho
                c.execute("UPDATE notas SET valor_restante = valor_restante + %s WHERE numero = %s",
                          (valor_empenho, numero_nota))
                c.execute("DELETE FROM empenhos WHERE id = %s", (empenho_id,))
                conn.commit()
            logging.info(f"Empenho {empenho_id} deletado com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao deletar empenho {empenho_id}: {e}")
            raise
        finally:
            self.release_connection(conn)

    def generate_excel_report(self):
        """Gera relatório em Excel com uma linha por empenho, incluindo a hierarquia."""
        notas = self.load_notas()
        empenhos = self.load_empenhos()
        naturezas = self.load_naturezas_despesa()
        planos = self.load_planos_internos()
        secoes = self.load_secoes_requisitantes()

        total_valor_geral = sum(n["valor"] for n in notas)
        total_restante_geral = sum(n["valor_restante"] for n in notas)
        total_empenhado_geral = sum(e["valor"] for e in empenhos)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório Detalhado de Empenhos"

        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

        headers = [
            "Plano Interno", "Natureza da Despesa", "PTRES", "Fonte",
            "Nº Nota", "Valor Original (R$)", "Valor Restante (R$)", "Descrição da Nota", "Prazo",
            "Data Empenho", "Valor Empenho (R$)", "Descrição Empenho", "Seção Requisitante"
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border

        for nota in notas:
            natureza = next((n for n in naturezas if n["codigo"] == nota["natureza_despesa_codigo"]), {"codigo": "N/A", "plano_interno_codigo": None})
            plano = next((p for p in planos if p["codigo"] == natureza["plano_interno_codigo"]), {"codigo": "N/A"}) if natureza["plano_interno_codigo"] else {"codigo": "N/A"}

            related_empenhos = [e for e in empenhos if e["numero_nota"] == nota["numero"]]
            if not related_empenhos:
                row_data = [
                    plano["codigo"], natureza["codigo"], nota["ptres_codigo"], nota["fonte_codigo"],
                    nota["numero"], nota["valor"], nota["valor_restante"], nota["descricao"], nota["prazo"],
                    "Nenhum empenho", "", "", "N/A"
                ]
                ws.append(row_data)
            else:
                for empenho in related_empenhos:
                    secao = next((s for s in secoes if s["codigo"] == empenho["secao_requisitante_codigo"]), {"codigo": "N/A"}) if empenho["secao_requisitante_codigo"] else {"codigo": "N/A"}
                    row_data = [
                        plano["codigo"], natureza["codigo"], nota["ptres_codigo"], nota["fonte_codigo"],
                        nota["numero"], nota["valor"], nota["valor_restante"], nota["descricao"], nota["prazo"],
                        empenho["data"], empenho["valor"], empenho["descricao"], secao["codigo"]
                    ]
                    ws.append(row_data)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = align_left if cell.column in [1, 2, 4, 8, 12, 13] else align_center

        ws.append([])  # Linha em branco
        total_row_idx = ws.max_row + 1
        ws.cell(row=total_row_idx, column=1, value="TOTAIS GERAIS").font = header_font
        ws.cell(row=total_row_idx, column=6, value=total_valor_geral).font = Font(bold=True)
        ws.cell(row=total_row_idx, column=7, value=total_restante_geral).font = Font(bold=True)
        ws.cell(row=total_row_idx, column=11, value=total_empenhado_geral).font = Font(bold=True)

        for col in range(1, len(headers) + 1):
            ws.cell(row=total_row_idx, column=col).border = border
            ws.cell(row=total_row_idx, column=col).alignment = align_center
        ws.cell(row=total_row_idx, column=1).fill = header_fill

        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col[0].column_letter].width = min(adjusted_width, 50)

        filename = "relatorios/relatorio_notas_credito.xlsx"
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        wb.save(filename)
        logging.info("Relatório Excel gerado com sucesso.")
        return filename

    def generate_pdf_report(self):
        """Gera relatório em PDF com uma linha por empenho, incluindo a hierarquia."""
        notas = self.load_notas()
        empenhos = self.load_empenhos()
        naturezas = self.load_naturezas_despesa()
        planos = self.load_planos_internos()
        secoes = self.load_secoes_requisitantes()

        total_valor_geral = sum(n["valor"] for n in notas)
        total_restante_geral = sum(n["valor_restante"] for n in notas)
        total_empenhado_geral = sum(e["valor"] for e in empenhos)

        filename = "relatorios/relatorio_notas_credito.pdf"
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        doc = SimpleDocTemplate(filename, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = []

        headers = [
            "Plano Interno", "Natureza da Despesa", "PTRES", "Fonte",
            "Nº Nota", "V. Original", "V. Restante", "Data Empenho", "V. Empenho", "Descrição Empenho", "Seção Requisitante"
        ]
        data = [headers]

        for nota in notas:
            natureza = next((n for n in naturezas if n["codigo"] == nota["natureza_despesa_codigo"]), {"codigo": "N/A", "plano_interno_codigo": None})
            plano = next((p for p in planos if p["codigo"] == natureza["plano_interno_codigo"]), {"codigo": "N/A"}) if natureza["plano_interno_codigo"] else {"codigo": "N/A"}

            related_empenhos = [e for e in empenhos if e["numero_nota"] == nota["numero"]]
            if not related_empenhos:
                data.append([
                    Paragraph(plano["codigo"], styles['BodyText']),
                    Paragraph(natureza["codigo"], styles['BodyText']),
                    Paragraph(nota["ptres_codigo"], styles['BodyText']),
                    Paragraph(nota["fonte_codigo"], styles['BodyText']),
                    Paragraph(nota["numero"], styles['BodyText']),
                    f"R$ {nota['valor']:.2f}", f"R$ {nota['valor_restante']:.2f}",
                    "Nenhum", "", Paragraph(nota["descricao"], styles['BodyText']), "N/A"
                ])
            else:
                for empenho in related_empenhos:
                    secao = next((s for s in secoes if s["codigo"] == empenho["secao_requisitante_codigo"]), {"codigo": "N/A"}) if empenho["secao_requisitante_codigo"] else {"codigo": "N/A"}
                    data.append([
                        Paragraph(plano["codigo"], styles['BodyText']),
                        Paragraph(natureza["codigo"], styles['BodyText']),
                        Paragraph(nota["ptres_codigo"], styles['BodyText']),
                        Paragraph(nota["fonte_codigo"], styles['BodyText']),
                        Paragraph(nota["numero"], styles['BodyText']),
                        f"R$ {nota['valor']:.2f}", f"R$ {nota['valor_restante']:.2f}",
                        empenho["data"], f"R$ {empenho['valor']:.2f}", Paragraph(empenho["descricao"], styles['BodyText']),
                        Paragraph(secao["codigo"], styles['BodyText'])
                    ])

        table = Table(data, colWidths=[80, 80, 80, 80, 80, 80, 80, 100, 80, 180, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E90FF")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ]))

        title = Paragraph("Relatório Detalhado de Notas de Crédito e Empenhos", styles['h1'])
        elements.append(title)
        elements.append(table)
        elements.append(Spacer(1, 20))

        total_data = [
            ["Totais Gerais", "", "", "", "Valor Original Total", "Valor Restante Total", "", "Valor Empenhado Total", "", ""],
            ["", "", "", "", f"R$ {total_valor_geral:.2f}", f"R$ {total_restante_geral:.2f}", "", f"R$ {total_empenhado_geral:.2f}", "", ""]
        ]
        total_table = Table(total_data, colWidths=[80, 80, 80, 80, 80, 80, 100, 80, 180, 80])
        total_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(total_table)

        doc.build(elements)
        logging.info("Relatório PDF gerado com sucesso.")
        return filename

# Configuração da Página Streamlit
st.set_page_config(page_title="Controle de Notas de Crédito", layout="wide")
st.title("Controle de Notas de Crédito")

# Inicialização do DataService
try:
    data_service = DataService()
except Exception as e:
    st.error(f"Erro ao inicializar a aplicação: {e}")
    st.stop()

# Funções de Validação
def validar_data(data_str):
    """Valida se uma string está no formato DD/MM/AAAA."""
    try:
        datetime.strptime(data_str, "%d/%m/%Y")
        return True
    except ValueError:
        return False

def validar_numerico(valor):
    """Valida se o valor é numérico."""
    try:
        return float(valor.replace(',', '.')) if valor else None
    except ValueError:
        return None

def validar_ptres(ptres):
    """Valida PTRES: exatamente 6 dígitos."""
    return ptres.isdigit() and len(ptres) == 6

def validar_fonte(fonte):
    """Valida Fonte: exatamente 10 dígitos."""
    return fonte.isdigit() and len(fonte) == 10

def validar_nota_numero(numero):
    """Valida Número da Nota: NC + 6 dígitos."""
    return bool(re.match(r"^NC\d{6}$", numero))

# Barra Lateral de Navegação
menu = [
    "🏠 Início",
    "📋 Adicionar Plano Interno",
    "📋 Adicionar Natureza da Despesa",
    "📋 Adicionar Seção Requisitante",
    "➕ Adicionar Nota",
    "📉 Registrar Empenho",
    "🗑️ Deletar Nota",
    "🗑️ Deletar Empenho",
    "📊 Visualizar Relatório",
    "📊 Empenhos por Seção",
    "📑 Relatório Excel",
    "📄 Relatório PDF"
]
opcao = st.sidebar.selectbox("Menu", menu)

# Estilo Personalizado
st.markdown("""
    <style>
    .stButton > button {
        background-color: #1E90FF;
        color: white;
        font-weight: bold;
        border-radius: 5px;
        padding: 10px;
        width: 100%;
    }
    .stButton > button:hover {
        background-color: #104E8B;
    }
    .stTextInput > div > input {
        background-color: #FFFFFF;
        color: black;
    }
    .stSelectbox > div > select {
        background-color: #FFFFFF;
        color: black;
    }
    </style>
""", unsafe_allow_html=True)

# Funções da Interface
if opcao == "🏠 Início":
    st.header("Bem-vindo ao Controle de Notas de Crédito")
    st.write("Use o menu lateral para gerenciar notas de crédito, empenhos e relatórios.")

elif opcao == "📋 Adicionar Plano Interno":
    st.header("Adicionar Plano Interno")
    with st.form("form_plano_interno"):
        codigo = st.text_input("Código (ex: PI001)")
        submit = st.form_submit_button("Salvar")
        if submit:
            try:
                if not codigo:
                    st.error("O campo código não pode estar vazio.")
                else:
                    data_service.save_plano_interno({"codigo": codigo.upper()})
                    st.success(f"Plano Interno {codigo} adicionado com sucesso!")
            except ValueError as e:
                st.error(str(e))
            except Exception as e:
                st.error(f"Ocorreu um erro: {e}")

elif opcao == "📋 Adicionar Natureza da Despesa":
    st.header("Adicionar Natureza da Despesa")
    planos = data_service.load_planos_internos()
    if not planos:
        st.warning("Nenhum plano interno cadastrado. Cadastre um plano interno primeiro.")
    else:
        with st.form("form_natureza_despesa"):
            plano_codigo = st.selectbox("Plano Interno", [p["codigo"] for p in planos])
            codigo = st.text_input("Código (ex: 3.3.90.39)")
            submit = st.form_submit_button("Salvar")
            if submit:
                try:
                    if not codigo:
                        st.error("O campo código não pode estar vazio.")
                    else:
                        data_service.save_natureza_despesa({
                            "codigo": codigo,
                            "plano_interno_codigo": plano_codigo
                        })
                        st.success(f"Natureza da Despesa {codigo} adicionada com sucesso!")
                except ValueError as e:
                    st.error(str(e))
                except Exception as e:
                    st.error(f"Ocorreu um erro: {e}")

elif opcao == "📋 Adicionar Seção Requisitante":
    st.header("Adicionar Seção Requisitante")
    with st.form("form_secao_requisitante"):
        codigo = st.text_input("Código (ex: SR001)")
        submit = st.form_submit_button("Salvar")
        if submit:
            try:
                if not codigo:
                    st.error("O campo código não pode estar vazio.")
                else:
                    data_service.save_secao_requisitante({"codigo": codigo.upper()})
                    st.success(f"Seção Requisitante {codigo} adicionada com sucesso!")
            except ValueError as e:
                st.error(str(e))
            except Exception as e:
                st.error(f"Ocorreu um erro: {e}")

elif opcao == "➕ Adicionar Nota":
    st.header("Adicionar Nota de Crédito")
    planos = data_service.load_planos_internos()
    naturezas = data_service.load_naturezas_despesa()
    if not planos or not naturezas:
        st.warning("Cadastre pelo menos um Plano Interno e uma Natureza da Despesa antes de adicionar uma nota.")
    else:
        with st.form("form_nota"):
            plano_codigo = st.selectbox("Plano Interno", [p["codigo"] for p in planos])
            naturezas_filtradas = [n["codigo"] for n in data_service.load_naturezas_despesa(plano_codigo)]
            natureza_codigo = st.selectbox("Natureza da Despesa", naturezas_filtradas)
            ptres_codigo = st.text_input("PTRES Código (6 dígitos)")
            fonte_codigo = st.text_input("Fonte Código (10 dígitos)")
            numero = st.text_input("Número da Nota (NC + 6 dígitos, ex: NC123456)")
            valor = st.text_input("Valor (ex: 1000.50)")
            descricao = st.text_input("Descrição")
            observacao = st.text_input("Observação (opcional)")
            prazo = st.text_input("Prazo Limite (DD/MM/AAAA)")
            submit = st.form_submit_button("Salvar")
            if submit:
                try:
                    if not all([plano_codigo, natureza_codigo, ptres_codigo, fonte_codigo, numero, valor, descricao, prazo]):
                        st.error("Preencha todos os campos obrigatórios.")
                    elif not validar_ptres(ptres_codigo):
                        st.error("O PTRES Código deve ter exatamente 6 dígitos.")
                    elif not validar_fonte(fonte_codigo):
                        st.error("O Fonte Código deve ter exatamente 10 dígitos.")
                    elif not validar_nota_numero(numero):
                        st.error("O Número da Nota deve ser NC seguido de 6 dígitos (ex: NC123456).")
                    elif not validar_data(prazo):
                        st.error("Prazo inválido! Use o formato DD/MM/AAAA (ex: 01/08/2025).")
                    else:
                        valor_float = validar_numerico(valor)
                        if valor_float is None or valor_float <= 0:
                            st.error("O valor deve ser um número positivo.")
                        else:
                            nota = {
                                "numero": numero.upper(),
                                "valor": valor_float,
                                "valor_restante": valor_float,
                                "descricao": descricao or "Sem descrição",
                                "observacao": observacao or "Sem observação",
                                "prazo": prazo,
                                "data_criacao": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                                "natureza_despesa_codigo": natureza_codigo,
                                "plano_interno_codigo": plano_codigo,
                                "ptres_codigo": ptres_codigo,
                                "fonte_codigo": fonte_codigo
                            }
                            data_service.save_nota(nota)
                            st.success(f"Nota {numero} adicionada com sucesso!")
                except ValueError as e:
                    st.error(str(e))
                except Exception as e:
                    st.error(f"Ocorreu um erro: {e}")

elif opcao == "📉 Registrar Empenho":
    st.header("Registrar Empenho")
    notas = data_service.load_notas()
    secoes = data_service.load_secoes_requisitantes()
    if not notas:
        st.warning("Nenhuma nota de crédito cadastrada para registrar um empenho.")
    elif not secoes:
        st.warning("Nenhuma seção requisitante cadastrada. Cadastre uma seção antes de registrar um empenho.")
    else:
        with st.form("form_empenho"):
            nota_opcoes = [f'{n["numero"]} (Saldo: R${n["valor_restante"]:.2f})' for n in notas]
            nota_selecionada = st.selectbox("Nota de Crédito", nota_opcoes)
            secao_codigo = st.selectbox("Seção Requisitante", [s["codigo"] for s in secoes])
            valor = st.text_input("Valor do Empenho")
            descricao = st.text_input("Descrição")
            submit = st.form_submit_button("Salvar")
            if submit:
                try:
                    numero_nota = nota_selecionada.split(" ")[0]
                    nota = next((n for n in notas if n["numero"] == numero_nota), None)
                    valor_float = validar_numerico(valor)
                    if not all([numero_nota, secao_codigo, valor, descricao]):
                        st.error("Preencha todos os campos.")
                    elif valor_float is None or valor_float <= 0:
                        st.error("O valor do empenho deve ser positivo.")
                    elif valor_float > nota["valor_restante"]:
                        st.error(f"Valor do empenho excede o saldo restante (R${nota['valor_restante']:.2f}).")
                    else:
                        nota["valor_restante"] -= valor_float
                        empenho = {
                            "numero_nota": numero_nota,
                            "valor": valor_float,
                            "descricao": descricao or "Empenho sem descrição",
                            "data": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                            "secao_requisitante_codigo": secao_codigo
                        }
                        data_service.save_empenho(empenho, nota)
                        st.success(f"Empenho de R${valor_float:.2f} registrado na nota {numero_nota}.")
                except ValueError as e:
                    st.error(str(e))
                except Exception as e:
                    st.error(f"Ocorreu um erro: {e}")

elif opcao == "🗑️ Deletar Nota":
    st.header("Deletar Nota de Crédito")
    notas = data_service.load_notas()
    if not notas:
        st.warning("Nenhuma nota de crédito cadastrada para deletar.")
    else:
        with st.form("form_deletar_nota"):
            nota_opcoes = [f'{n["numero"]} (Saldo: R${n["valor_restante"]:.2f})' for n in notas]
            nota_selecionada = st.selectbox("Nota de Crédito", nota_opcoes)
            submit = st.form_submit_button("Deletar")
            if submit:
                try:
                    numero_nota = nota_selecionada.split(" ")[0]
                    if st.checkbox("Confirmar exclusão (todos os empenhos associados serão excluídos)"):
                        data_service.delete_nota(numero_nota)
                        st.success(f"Nota {numero_nota} deletada com sucesso!")
                    else:
                        st.warning("Marque a caixa de confirmação para deletar.")
                except ValueError as e:
                    st.error(str(e))
                except Exception as e:
                    st.error(f"Ocorreu um erro: {e}")

elif opcao == "🗑️ Deletar Empenho":
    st.header("Deletar Empenho")
    empenhos = data_service.load_empenhos()
    if not empenhos:
        st.warning("Nenhum empenho cadastrado para deletar.")
    else:
        with st.form("form_deletar_empenho"):
            empenho_opcoes = [f'ID {e["id"]} - Nota {e["numero_nota"]} (Valor: R${e["valor"]:.2f}, Data: {e["data"]}, Seção: {e["secao_requisitante_codigo"]})' for e in empenhos]
            empenho_selecionado = st.selectbox("Empenho", empenho_opcoes)
            submit = st.form_submit_button("Deletar")
            if submit:
                try:
                    empenho_id = int(empenho_selecionado.split(" ")[1])
                    if st.checkbox("Confirmar exclusão"):
                        data_service.delete_empenho(empenho_id)
                        st.success(f"Empenho ID {empenho_id} deletado com sucesso!")
                    else:
                        st.warning("Marque a caixa de confirmação para deletar.")
                except ValueError as e:
                    st.error(str(e))
                except Exception as e:
                    st.error(f"Ocorreu um erro: {e}")

elif opcao == "📊 Visualizar Relatório":
    st.header("Relatório Detalhado")
    notas = data_service.load_notas()
    empenhos = data_service.load_empenhos()
    naturezas = data_service.load_naturezas_despesa()
    planos = data_service.load_planos_internos()
    secoes = data_service.load_secoes_requisitantes()

    data = []
    for nota in notas:
        natureza = next((n for n in naturezas if n["codigo"] == nota["natureza_despesa_codigo"]), {"codigo": "N/A", "plano_interno_codigo": None})
        plano = next((p for p in planos if p["codigo"] == natureza["plano_interno_codigo"]), {"codigo": "N/A"}) if natureza["plano_interno_codigo"] else {"codigo": "N/A"}
        related_empenhos = [e for e in empenhos if e["numero_nota"] == nota["numero"]]
        if not related_empenhos:
            data.append({
                "Plano Interno": plano["codigo"],
                "Natureza da Despesa": natureza["codigo"],
                "PTRES": nota["ptres_codigo"],
                "Fonte": nota["fonte_codigo"],
                "Nº Nota": nota["numero"],
                "V. Original": f"R${nota['valor']:.2f}",
                "V. Restante": f"R${nota['valor_restante']:.2f}",
                "Data Empenho": "Nenhum",
                "V. Empenho": "",
                "Descrição Empenho": nota["descricao"],
                "Seção Requisitante": "N/A"
            })
        else:
            for empenho in related_empenhos:
                secao = next((s for s in secoes if s["codigo"] == empenho["secao_requisitante_codigo"]), {"codigo": "N/A"}) if empenho["secao_requisitante_codigo"] else {"codigo": "N/A"}
                data.append({
                    "Plano Interno": plano["codigo"],
                    "Natureza da Despesa": natureza["codigo"],
                    "PTRES": nota["ptres_codigo"],
                    "Fonte": nota["fonte_codigo"],
                    "Nº Nota": nota["numero"],
                    "V. Original": f"R${nota['valor']:.2f}",
                    "V. Restante": f"R${nota['valor_restante']:.2f}",
                    "Data Empenho": empenho["data"],
                    "V. Empenho": f"R${empenho['valor']:.2f}",
                    "Descrição Empenho": empenho["descricao"],
                    "Seção Requisitante": secao["codigo"]
                })

    if data:
        df = pd.DataFrame(data)
        st.dataframe(df, use_container_width=True)
    else:
        st.info("Nenhum dado disponível para exibir.")

elif opcao == "📊 Empenhos por Seção":
    st.header("Empenhos por Seção Requisitante")
    secoes = data_service.load_secoes_requisitantes()
    if not secoes:
        st.warning("Nenhuma seção requisitante cadastrada.")
    else:
        secao_codigo = st.selectbox("Seção Requisitante", [s["codigo"] for s in secoes])
        if secao_codigo:
            empenhos = data_service.load_empenhos()
            data = [
                {
                    "Nº Nota": e["numero_nota"],
                    "Valor Empenho": f"R${e['valor']:.2f}",
                    "Data Empenho": e["data"],
                    "Descrição Empenho": e["descricao"]
                }
                for e in empenhos if e["secao_requisitante_codigo"] == secao_codigo
            ]
            if data:
                df = pd.DataFrame(data)
                st.dataframe(df, use_container_width=True)
            else:
                st.info("Nenhum empenho encontrado para esta seção.")

elif opcao == "📑 Relatório Excel":
    st.header("Gerar Relatório Excel")
    if st.button("Gerar Relatório"):
        try:
            filename = data_service.generate_excel_report()
            with open(filename, "rb") as f:
                st.download_button(
                    label="Baixar Relatório Excel",
                    data=f,
                    file_name="relatorio_notas_credito.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            st.success("Relatório Excel gerado com sucesso!")
        except Exception as e:
            st.error(f"Erro ao gerar relatório: {e}")

elif opcao == "📄 Relatório PDF":
    st.header("Gerar Relatório PDF")
    if st.button("Gerar Relatório"):
        try:
            filename = data_service.generate_pdf_report()
            with open(filename, "rb") as f:
                st.download_button(
                    label="Baixar Relatório PDF",
                    data=f,
                    file_name="relatorio_notas_credito.pdf",
                    mime="application/pdf"
                )
            st.success("Relatório PDF gerado com sucesso!")
        except Exception as e:
            st.error(f"Erro ao gerar relatório: {e}")